import sys
from scipy.stats import beta,chi2,weibull_min,lognorm,gamma
import math
import statistics
import xlwt
import numpy


print("Alfa ve Beta paremetreleri sizden rastgele GAMMA dağılım değeri üretmek bizden")
ax = float(input("alfa değerini giriniz: "))

print("*" * 50, end="\n\n\n")


def hesap(ax):
    z, k, m, n, p, t = 0, 0, 0, 0, 0, 0
    studentGuvenAlt, aadmGuvenAlt, maadGuvenAlt, madmGuvenAlt, johnsonGuvenAlt, chenGuvenAlt = list(), list(), list(), list(), list(), list()
    studentGuvenUst, aadmGuvenUst, maadGuvenUst, madmGuvenUst, johnsonGuvenUst, chenGuvenUst = list(), list(), list(), list(), list(), list()

    workbook = xlwt.Workbook()
    sayfa = workbook.add_sheet("Sayfa1")
    sayfa.write(0, 1, "Student-t")
    sayfa.write(0, 3, "AADM-t")
    sayfa.write(0, 5, "MAAD-t")
    sayfa.write(0, 7, "MADM-t")
    sayfa.write(0, 9, "Johnson-t")
    sayfa.write(0, 11, "Chen-t")

    for item in range(0, 13):
        if item == 0:
            sayfa.write(1, 0, "n")
        elif item % 2 == 0:
            sayfa.write(1, item, "AW")

        else:
            sayfa.write(1, item, "CP")

    for i in range(5, 10):
        for j in range(1, 2501):

            x = chi2.rvs(ax, size=i)


            meanx = round(statistics.mean(x), 4)
            medianx = round(statistics.median(x), 4)
            stdevx = round(statistics.stdev(x), 4)
            aadmx = round((math.sqrt(math.pi / 2) / i) * sum(abs(x - medianx)), 4)
            maadx = round(statistics.median(abs(x - meanx)), 4)
            madmx = round(statistics.median(abs(x - medianx)), 4)

            toplam = 0
            for k in range(0, i):
                toplam = toplam + ((x[k] - meanx) ** 3)

            m3 = (i / ((i - 1) * (i - 2))) * toplam


            studentalt = round(meanx - cell[i - 5] * stdevx / math.sqrt(i), 4)
            studentust = round(meanx + cell[i - 5] * stdevx / math.sqrt(i), 4)
            aadmalt = round(meanx - cell[i - 5] * aadmx / math.sqrt(i), 4)
            aadmust = round(meanx + cell[i - 5] * aadmx / math.sqrt(i), 4)
            maadalt = round(meanx - cell[i - 5] * maadx / math.sqrt(i), 4)
            maadust = round(meanx + cell[i - 5] * maadx / math.sqrt(i), 4)
            madmalt = round(meanx - cell[i - 5] * madmx / math.sqrt(i), 4)
            madmust = round(meanx + cell[i - 5] * madmx / math.sqrt(i), 4)
            johnsonalt=round((meanx+(m3/(6*i*(stdevx**2))))-cell[i-5]*math.sqrt(i)*stdevx,4)
            johnsonust=round((meanx+(m3/(6*i*(stdevx**2))))+cell[i-5]*math.sqrt(i)*stdevx,4)
            chenalt = round(meanx - (
                        cell[i - 5] + (((m3 / (stdevx ** 3)) * (1 + 2 * (cell[i - 5] ** 2))) / (6 * math.sqrt(i))) + (
                            (((m3 / (stdevx ** 3)) ** 2) * (cell[i - 5] + 2 * (cell[i - 5]) ** 2) / 9 * i)) + math.sqrt(
                    i) * stdevx))
            chenust = round(meanx + (
                        cell[i - 5] + (((m3 / (stdevx ** 3)) * (1 + 2 * (cell[i - 5] ** 2))) / (6 * math.sqrt(i))) + (
                             (((m3 / (stdevx ** 3)) ** 2) * (cell[i - 5] + 2 * (cell[i - 5]) ** 2) / 9 * i)) + math.sqrt(
                    i) * stdevx))

            studentGuvenAlt.append(studentalt)
            studentGuvenUst.append(studentust)
            aadmGuvenAlt.append(aadmalt)
            aadmGuvenUst.append(aadmust)
            maadGuvenAlt.append(maadalt)
            maadGuvenUst.append(maadust)
            madmGuvenAlt.append(madmalt)
            madmGuvenUst.append(madmust)
            johnsonGuvenAlt.append(johnsonalt)
            johnsonGuvenUst.append(johnsonust)
            chenGuvenAlt.append(chenalt)
            chenGuvenUst.append(chenust)

            if studentalt <= ax-1 <= studentust:
                z = z + 1

            if aadmalt <= ax-1 <= aadmust:
                k = k + 1

            if madmalt <= ax-1 <= madmust:
                m = m + 1

            if maadalt <= ax-1 <= maadust:
                n = n + 1

            if johnsonalt <= ax - 1 <= johnsonust:
                p = p + 1

            if chenalt <= ax - 1 <= chenust:
                t = t + 1

        sayfa.write(i - 3, 0, f"{i}")
        sayfa.write(i - 3, 1, f"{round(z / 2500, 4)}")
        sayfa.write(i - 3, 2,
                    f"{round(statistics.mean(studentGuvenUst) - statistics.mean(studentGuvenAlt), 4)}")
        sayfa.write(i - 3, 3, f"{round(k / 2500, 4)}")
        sayfa.write(i - 3, 4, f"{round(statistics.mean(aadmGuvenUst) - statistics.mean(aadmGuvenAlt), 4)}")
        sayfa.write(i - 3, 5, f"{round(n / 2500, 4)}")
        sayfa.write(i - 3, 6, f"{round(statistics.mean(maadGuvenUst) - statistics.mean(maadGuvenAlt), 4)}")
        sayfa.write(i - 3, 7, f"{round(m / 2500, 4)}")
        sayfa.write(i - 3, 8, f"{round(statistics.mean(madmGuvenUst) - statistics.mean(madmGuvenAlt), 4)}")
        sayfa.write(i - 3, 9, f"{round(p / 2500, 4)}")
        sayfa.write(i - 3, 10, f"{round(statistics.mean(johnsonGuvenUst) - statistics.mean(johnsonGuvenAlt), 4)}")
        sayfa.write(i - 3, 11, f"{round(t / 2500, 4)}")
        sayfa.write(i - 3, 12, f"{round(statistics.mean(chenGuvenUst) - statistics.mean(chenGuvenAlt), 4)}")

        workbook.save(f'K({ax} ).xls')  # excelisim

        z, k, m, n, p, t = 0, 0, 0, 0, 0, 0
        studentGuvenAlt, aadmGuvenAlt, maadGuvenAlt, madmGuvenAlt, johnsonGuvenAlt, chenGuvenAlt = list(), list(), list(), list(), list(), list()
        studentGuvenUst, aadmGuvenUst, maadGuvenUst, madmGuvenUst, johnsonGuvenUst, chenGuvenUst = list(), list(), list(), list(), list(), list()


if __name__ == "__main__":
    import openpyxl as xl

    wb = xl.load_workbook("tTablosu.xlsx")
    sayfax = wb["Sayfa1"]
    cell = list()
    for i in range(7, 150):
        if i < 34:
            cell.append(sayfax.cell(i, 7).value / 1000)
        else:
            cell.append(2)

hesap(ax)

sys.exit()